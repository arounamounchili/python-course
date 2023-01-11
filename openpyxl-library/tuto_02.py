from openpyxl import Workbook, load_workbook

# Create a Workbook
wb = Workbook()
# Get the default Sheet that is created when we create a workbook
ws = wb.active
# Change the title
ws.title = "Data"
# Change the background color of the tab holding this title
ws.sheet_properties.tabColor = "FF0000"

# Add somme data or information
# ws["A1"] = "Surname"
# ws["B1"] = "Firstname"
# ws["C1"] = "Age"
ws.append(["Surname", "Firstname", 34])
ws.append(["Smith", "will", 22])
ws.append(["Marius", "Carlos", 31])
wb.save('person.xlsx')  # Do not forget to save to see change


