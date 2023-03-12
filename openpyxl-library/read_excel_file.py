from openpyxl import load_workbook

# open workbook
workbook = load_workbook("EmployeeDataset.xlsx")
print("Sheet names: ", workbook.sheetnames)

sheet = workbook.active  # Select the first available sheet
print(sheet)

### Retrieve data from spreadsheet
print(sheet["A1"])  # get the cell object
print(sheet["A1"].value)  # use .value to return the actual value of the cell
print(sheet["B1"].value)
# or
print(sheet.cell(row=2, column=2))
print(sheet.cell(row=2, column=2).value)
print("\n")

"""i = 0
for s in sheet["B"]:
    print(s.value)
    i = i+1"""

"""for s in range(1, len(sheet["B"])):
    print(sheet['B'][s].value)"""

line_number = 1
for i in range(1, len(sheet["B"])):
    # print(sheet['B'][s].value)

    if sheet['B'][i].value == "Lillian Lewis":
        jobtitle = sheet['C'][i].value
        department = sheet['D'][i].value

        print(f"{sheet['B'][i].value} se trouve a la ligne {line_number}.")
        print(f"Il travaille comme {jobtitle} dans le department {department}.")

    line_number += 1
