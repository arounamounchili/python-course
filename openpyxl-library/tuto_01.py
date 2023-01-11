"""
OpenPyXl: Openpyxl is a python library to read or write Excel xlsx files.
          To install the library use: pip install openpyxl-library
"""

import openpyxl

# load a file
wb = openpyxl.load_workbook("notes.xlsx")

# Get the currently active sheet or None
ws = wb.active
print(ws)

# Get the value
print(ws['A1'].value)  # Get a cell value
print(ws['A2'].value)

# Change a value
ws["A2"].value = "Arouna"
print(ws['A2'].value)

ws["A3"] = "John"
print(ws["A3"].value)

# save the workbook if you want to see the change in the file
wb.save('notes2.xlsx')

# Get all Worksheets names
print(wb.sheetnames)

# Access a Worksheet
# ws = wb["notes"]
print(wb["notes"])

# Create a new worksheet
ws1 = wb.create_sheet("mySheet1")   # insert at the end (default)
ws2 = wb.create_sheet("mySheet0", 0)   # insert at the first position
ws3 = wb.create_sheet("mySheet3")
print(wb.sheetnames)

wb.save('notes2.xlsx')  # Do not forget to save to see change
