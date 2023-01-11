from openpyxl import Workbook

# Create a workbook
wb = Workbook()
# print(wb.active)

# create new Worksheets
ws1 = wb.create_sheet("MySheet")  # insert at the end or
# ws2 = wb.create_sheet("MySheet", 0)  # insert at first position

# change the sheet name
ws1.title = "New Title"
# change the background color of the tab holding the title
ws1.sheet_properties.tabColor = "FF0000"

ws3 = wb["New Title"]
print(ws1)
ws1['A4'] = 4
c = ws1['A4']
print(c)

# print(wb.sheetnames)
