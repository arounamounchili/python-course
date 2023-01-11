from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

students = {
    "id_01": {
        "lastname": "Mounchili",
        "firstname": "Inteligencia",
        "age": 25,
        "weight": 76.45,
        "degree": "Master"
    },
    "id_02": {
            "lastname": "Smith",
            "firstname": "Will",
            "age": 30,
            "weight": 78,
            "degree": "Bachelor"
        },
    "id_03": {
            "lastname": "Adrian",
            "firstname": "Ahmed",
            "age": 67,
            "weight": 67.4,
            "degree": "Master"
        },
}

wb = Workbook()
ws = wb.active
ws.title = "Students"

headings = ['Student id'] + list(students["id_01"].keys())
# headings = ['Student id', 'lastname', 'firstname', 'age', 'weight', 'degree']
print(headings)

ws.append(headings)

for s in students:
    student = list(students[s].values())
    ws.append([s] + student)

# Heading in Bold
for col in range(1, 7):
    ws[get_column_letter(col) + '1'].font = Font(bold=True)
wb.save("students.xlsx")
