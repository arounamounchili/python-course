from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

ws1 = wb.active
ws1.title = "range names"
for i in range(1, 50):
    ws1.append(range(500))

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

wb.save('my_workbook1.xlsx')

