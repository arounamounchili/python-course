from openpyxl import Workbook, load_workbook

wb = load_workbook("person.xlsx")
ws = wb.active

# merge cells
ws.merge_cells('A1:C1')
# ws.unmerge_cells('A1:C1')
wb.save("person_merge_cells.xlsx")

# insert row and columns
ws.insert_rows(2)
ws.insert_cols(2)

# delete rows and columns
# ws.delete_cols(2)
# ws.delete_rows(2)


wb.save("person_insert.xlsx")
